from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook


    
		
import tkinter as tk

#################################################################3

# Openpyxl searching data functions
# reference: 
# get worksheet by name: https://stackoverflow.com/questions/36814050/openpyxl-get-sheet-by-name

# code source: https://stackoverflow.com/questions/50491839/python-openpyxl-find-strings-in-column-and-return-row-number
# wb = load_workbook(filename = 'empty_book.xlsx')
# ws = wb.active
# col_idx, row = search_value_in_col_index(ws, "_my_search_string_")

def search_value_in_column(ws, search_string, column="A"):    
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)        
        if ws[coordinate].value == search_string:             
            return column, row   
    return column, None


def search_value_in_col_idx(ws, search_string, col_idx=1):
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx, None


def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row 
def only_compare_date(datetime_a, datetime_b):
    '''
    if type not the same return false
    if the type matched, compare only the date
    '''
    if type(datetime_a)==type(datetime_b):
        return datetime_a.date() == datetime_b.date()
    else:
        return False

def search_value_all_records_in_column(ws, match_value, column="A",compare_func=None):  
    '''
    This function return a list of rows in the column that the value matched
    compare_func is in case that we need to ONLY compare part of the value such as only date and does not care the time
    '''
    row_list = []
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)          
        if compare_func(ws[coordinate].value,match_value):             
            row_list.append(row)   
    return row_list

def get_last_new_row(ws):
    '''
    column = "A"
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)        
        if ws[coordinate].value == None: 
            print("value = ")
            print(ws[coordinate].value)
            break
    return row
    WE ASSUME: There is no empty row in the database
    '''
    column,row =  search_value_in_column(ws, None, "A")
    if row==None:
        return ws.max_row+1
    return row

#######################################3333
# WorkTime

class WorkTime:
    def __init__(self, filename="hello_world.xlsx"):
        # Openpyxl
        self.data_filename = filename
        self.worker_table_name="workers"
        self.data_wb = load_workbook(filename = self.data_filename)
        
        
        #UI
        self.worktime_ui = tk.Tk()
        self.employee_num=tk.StringVar()
        self.msg=tk.StringVar()
        self.worktime_ui.geometry("720x300")
        self.worktime_ui.title("上班打卡")
        self.employee_num_lb=tk.Label(self.worktime_ui,text="請輸入員工工號:")
        self.employee_num_lb.pack()
        self.employee_entry = tk.Entry(self.worktime_ui,textvariable=self.employee_num)
        self.employee_entry.pack()
        self.login_btn = tk.Button(self.worktime_ui,text="登入",command=self.register)
        self.login_btn.pack()
        self.result_msg=tk.Label(self.worktime_ui,fg="red",textvariable=self.msg)
        self.result_msg.pack()
        self.worktime_ui.mainloop()
    def get_user_name(self, employee_num): 
        '''
            Open spreadsheet workers (ID, name) and search the matched user name
            return user name
            
        '''                
        workers_ws = self.data_wb[self.worker_table_name]        
        column, row = search_value_in_column(workers_ws, employee_num, column="A")
              
        if row == None: 
            # not found
            return None       
        coordinate = "{}{}".format("B", row)        
        return workers_ws[coordinate].value
        
    def is_login_history_normal(self):
        '''
            exist 2 or times of 2 login records 
        '''
    def build_log_record(self):
        '''
            is_login_history?
            Correct -> create and save log record
            Wrong -> create and save log record 
                     Show error message that need to contact with manager.
        '''
    def save_sp(self):
        self.data_wb.save(self.data_filename)
        
    def is_no_open_log(self):
        return True
        
    def save_log_record2sp(self,user_name,current_date_time):
        '''
        
        1. Search all todays log and add the new check in time
                 If today's records are all compleate log => checkin => create a new log with check in.
                 If there are one open record (has check in but no checkout) => checkout => put the log in the last today record
        2. alert log error if there are
        '''
        employee_ws = self.data_wb[user_name]   
        
        
        ###### search today's record. 
        date_log_record_list = search_value_all_records_in_column(employee_ws, current_date_time, "B",only_compare_date)
        #  if no today record 
        #      save ABC 
        #  else 
        #      D is empty  Save D##################################
        print("dates are ")
        print(date_log_record_list)
        # check the login should be checkout or checkin
        is_checkout = False
        for row in date_log_record_list:
            # if column D is empty(None) => No checkout, this is this time should be checkout 
            coordinate = "{}{}".format("D", row)         
            print ("row:", row , "column D is ", employee_ws[coordinate].value )
            if (employee_ws[coordinate].value == None):
                employee_ws[coordinate] = current_date_time.time()
                is_checkout = True
        
        if is_checkout == False:
            # create a new row
            row = get_last_new_row(employee_ws)
            ################# Save ABC #############333333
            # name
            coordinate = "{}{}".format("A", row) 
            employee_ws[coordinate] = user_name
            # checking there is no WRONG checkout before
            # check in time
            coordinate = "{}{}".format("B", row) 
            employee_ws[coordinate] = current_date_time.date()
            #Important Note: in spreadsheet, there is only datetime data. The date or time is depends on VIEW
            # therefore, when you save date() it make the time to be 00:00:00
            # This is important for compare the date info

            coordinate = "{}{}".format("C", row) 
            employee_ws[coordinate] = current_date_time.time()
            ############# save D sta#####################3
            coordinate = "{}{}".format("D", row) 
            employee_ws[coordinate] = current_date_time.time()
        
        self.save_sp()
        
    def check_pw(self):


        now = datetime.now()
        
        current_time = now.strftime("%m/%d -- %H:%M:%S")
        #print("Current Time =", current_time)
        #https://stackoverflow.com/questions/50491839/python-openpyxl-find-strings-in-column-and-return-row-number
        user_id = self.employee_num.get()
        user_name = self.get_user_name(user_id)
        if user_name == None: 
            user_name = "user not Found"        
        msg_str = "歡迎 "+user_name+" 登入,您簽到的時間是:"+current_time
        self.msg.set(msg_str)
        return user_name, now
 

        #workbook.save(self.data_filename)
    def register(self):
        user_name, now = self.check_pw()
         ############################333 
         
        self.save_log_record2sp(user_name,now)
        
         
############################################33
######## End of work time
##########################################
def main():
    
    wt = WorkTime(filename="work_log.xlsx")    

    
    

if __name__ == "__main__":
    main()
    
    
    
    


