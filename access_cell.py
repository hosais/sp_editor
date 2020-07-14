# this is an example to access one cell and print its value on the screen

#from openpyxl import load_workbook
#wb2 = load_workbook('test.xlsx')
#print(wb2.sheetnames)
from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")